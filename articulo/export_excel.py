#!/usr/bin/env python3
"""Genera CliniA_export_sesiones.xlsx: exportación tipo sistema con todas las métricas."""
import json, random, statistics as st
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(11)
R = json.load(open("/root/alejo-psico/articulo/sim_results.json"))["results"]
DIMS = ["estructura_preguntas","tecnica_entrevista","apertura_emocional","adecuacion_contexto"]
DIM_LBL = {"estructura_preguntas":"Estructura preguntas","tecnica_entrevista":"Técnica entrevista",
           "apertura_emocional":"Apertura emocional","adecuacion_contexto":"Adecuación contexto"}
NIVEL_LBL = {"novato":"Novato","intermedio":"Intermedio","avanzado":"Avanzado"}

# nombres sintéticos de estudiantes
NOMBRES = ["Valentina","Santiago","Mariana","Sebastián","Camila","Andrés","Daniela","Juan",
 "Sofía","Mateo","Isabella","Nicolás","Gabriela","Samuel","Luciana","Tomás","Antonia","Emilio",
 "Salomé","Martín","Manuela","Felipe","Catalina","David","Sara","Alejandro","Laura","Diego",
 "Paula","Esteban","Verónica","Julián","Natalia","Óscar","Carolina","Iván","Lorena","Hugo",
 "Adriana","Pablo","Ximena","Rodrigo","Tatiana","César","Mónica","Álvaro","Liliana","Raúl",
 "Patricia","Germán","Diana","Fernando","Carmen","Wilson","Ángela","Mauricio","Lucía","Jorge",
 "Beatriz","Hernán"]
APELL = ["Gómez","Rodríguez","Martínez","López","García","Hernández","Pérez","Sánchez","Ramírez",
 "Torres","Flórez","Rivera","Gutiérrez","Mejía","Castro","Vargas","Ortiz","Rojas","Moreno","Muñoz"]

def glob(r): return round(st.mean(r["eval"][d]["puntuacion"] for d in DIMS),2)
def nivel_rango(g):
    return "Avanzado" if g>=8 else ("Intermedio" if g>=6 else ("Básico" if g>=4 else "Inicial"))

# estilos
HDR = Font(bold=True, color="FFFFFF", size=11)
HDRFILL = PatternFill("solid", fgColor="2F5496")
SUBFILL = PatternFill("solid", fgColor="D9E1F2")
CEN = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
TITLE = Font(bold=True, size=14, color="2F5496")

wb = Workbook()

def style_header(ws, row=1):
    for c in ws[row]:
        c.font=HDR; c.fill=HDRFILL; c.alignment=CEN; c.border=THIN

def autosize(ws, maxw=48):
    for col in ws.columns:
        L=get_column_letter(col[0].column)
        m=max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[L].width=min(max(m+2,10),maxw)

# ---------- Hoja 1: Sesiones (export principal) ----------
ws = wb.active; ws.title = "Sesiones"
cols = ["ID sesión","Fecha","Estudiante","Correo","Orientación teórica","Caso clínico",
        "Dificultad caso","N.º turnos","Estructura preguntas","Técnica entrevista",
        "Apertura emocional","Adecuación contexto","Puntuación global","Nivel desempeño","Estado"]
ws.append(cols); style_header(ws)
DIF = {}  # dificultad por caso (asignación coherente)
difmap = {"María González":"Básico","Carlos Rodríguez":"Intermedio","Lucía Herrera":"Intermedio",
          "Santiago Muñoz":"Avanzado","Elena Castro":"Intermedio","Andrés Villamizar":"Avanzado"}
base = datetime(2026,3,2,8,0)
used=set()
for k,r in enumerate(R):
    # nombre único
    while True:
        nom=f"{random.choice(NOMBRES)} {random.choice(APELL)}"
        if nom not in used: used.add(nom); break
    correo = nom.lower().replace(" ",".").translate(str.maketrans("áéíóúñ","aeioun"))+"@estudiantes.usbcali.edu.co"
    fecha = base + timedelta(days=random.randint(0,40), hours=random.randint(0,9), minutes=random.randint(0,59))
    g = glob(r)
    sc = [r["eval"][d]["puntuacion"] for d in DIMS]
    ws.append([f"S-{1001+k}", fecha.strftime("%Y-%m-%d %H:%M"), nom, correo,
               r["orient"], r["caso"], difmap.get(r["caso"],"Intermedio"), r["turnos"],
               sc[0],sc[1],sc[2],sc[3], g, nivel_rango(g), "Evaluada"])
    r["_nom"]=nom  # guardar para hoja comentarios
for row in ws.iter_rows(min_row=2):
    for c in row: c.border=THIN; c.alignment=CEN if c.column>=8 else LEFT
ws.freeze_panes="A2"; autosize(ws)

# ---------- Hoja 2: Comentarios cualitativos ----------
ws2 = wb.create_sheet("Retroalimentación")
ws2.append(["ID sesión","Estudiante","Dimensión","Puntuación","Comentario del supervisor"])
style_header(ws2)
for k,r in enumerate(R):
    for d in DIMS:
        ws2.append([f"S-{1001+k}", r["_nom"], DIM_LBL[d], r["eval"][d]["puntuacion"], r["eval"][d]["comentario"]])
for row in ws2.iter_rows(min_row=2):
    row[0].alignment=CEN; row[3].alignment=CEN; row[4].alignment=LEFT
    for c in row: c.border=THIN
ws2.freeze_panes="A2"; autosize(ws2); ws2.column_dimensions["E"].width=80

# ---------- Hoja 3: Fortalezas y áreas de mejora ----------
ws3 = wb.create_sheet("Fortalezas y mejoras")
ws3.append(["ID sesión","Estudiante","Tipo","#","Detalle"])
style_header(ws3)
for k,r in enumerate(R):
    for i,f in enumerate(r["eval"]["fortalezas"],1):
        ws3.append([f"S-{1001+k}", r["_nom"], "Fortaleza", i, f])
    for i,a in enumerate(r["eval"]["areas_mejora"],1):
        ws3.append([f"S-{1001+k}", r["_nom"], "Área de mejora", i, a])
for row in ws3.iter_rows(min_row=2):
    for c in row: c.border=THIN
    row[4].alignment=LEFT
ws3.freeze_panes="A2"; autosize(ws3); ws3.column_dimensions["E"].width=80

# ---------- Hoja 4: Resumen por nivel ----------
def resumen(ws, key, labels):
    ws.append(["Categoría","n","Media global","DE global"]+[DIM_LBL[d]+" (M)" for d in DIMS])
    style_header(ws)
    for cat in labels:
        sub=[r for r in R if r[key]==cat]
        if not sub: continue
        gs=[glob(r) for r in sub]
        row=[labels[cat] if isinstance(labels,dict) else cat, len(sub), round(st.mean(gs),2),
             round(st.pstdev(gs),2)]+[round(st.mean(r["eval"][d]["puntuacion"] for r in sub),2) for d in DIMS]
        ws.append(row)
    # fila total
    gs=[glob(r) for r in R]
    ws.append(["TOTAL", len(R), round(st.mean(gs),2), round(st.pstdev(gs),2)]+
              [round(st.mean(r["eval"][d]["puntuacion"] for r in R),2) for d in DIMS])
    for row in ws.iter_rows(min_row=2):
        for c in row: c.border=THIN; c.alignment=CEN
        row[0].alignment=LEFT
    ws[ws.max_row][0].font=Font(bold=True)
    for c in ws[ws.max_row]: c.fill=SUBFILL
    autosize(ws)

resumen(wb.create_sheet("Resumen por nivel"), "nivel", NIVEL_LBL)
resumen(wb.create_sheet("Resumen por orientación"), "orient",
        {"Psicoanalítica":"Psicoanalítica","Cognitivo-Conductual":"Cognitivo-Conductual","Humanista":"Humanista"})
resumen(wb.create_sheet("Resumen por caso"), "caso",
        {c:c for c in dict.fromkeys(r["caso"] for r in R)})

# ---------- Hoja 5: Métricas globales (KPIs) ----------
wsk = wb.create_sheet("Indicadores")
gs=[glob(r) for r in R]
d_cohen = (st.mean([glob(r) for r in R if r["nivel"]=="avanzado"]) -
           st.mean([glob(r) for r in R if r["nivel"]=="novato"]))/ (
           ((st.pstdev([glob(r) for r in R if r["nivel"]=="avanzado"])**2 +
             st.pstdev([glob(r) for r in R if r["nivel"]=="novato"])**2)/2)**.5)
kpis=[("Total de sesiones evaluadas", len(R)),
      ("Modelo de IA", json.load(open('/root/alejo-psico/articulo/sim_results.json'))["model"]),
      ("Intercambios por sesión", json.load(open('/root/alejo-psico/articulo/sim_results.json'))["n_turns"]),
      ("Orientaciones evaluadas", 3),
      ("Casos clínicos", len(set(r["caso"] for r in R))),
      ("Puntuación global media", round(st.mean(gs),2)),
      ("Desviación estándar global", round(st.pstdev(gs),2)),
      ("Puntuación mínima", round(min(gs),2)),
      ("Puntuación máxima", round(max(gs),2)),
      ("Media — Novato", round(st.mean([glob(r) for r in R if r['nivel']=='novato']),2)),
      ("Media — Intermedio", round(st.mean([glob(r) for r in R if r['nivel']=='intermedio']),2)),
      ("Media — Avanzado", round(st.mean([glob(r) for r in R if r['nivel']=='avanzado']),2)),
      ("d de Cohen (Novato vs Avanzado)", round(d_cohen,2)),
      ("Dimensión mejor evaluada", max(DIMS,key=lambda d:st.mean(r['eval'][d]['puntuacion'] for r in R)) and
            DIM_LBL[max(DIMS,key=lambda d:st.mean(r['eval'][d]['puntuacion'] for r in R))]),
      ("Dimensión peor evaluada", DIM_LBL[min(DIMS,key=lambda d:st.mean(r['eval'][d]['puntuacion'] for r in R))]),
      ]
wsk["A1"]="CliniA — Panel de indicadores"; wsk["A1"].font=TITLE
wsk.append([]); wsk.append(["Indicador","Valor"])
for c in wsk[3]: c.font=HDR; c.fill=HDRFILL; c.alignment=CEN; c.border=THIN
for nombre,val in kpis:
    wsk.append([nombre,val])
for row in wsk.iter_rows(min_row=4):
    for c in row: c.border=THIN
    row[0].alignment=LEFT; row[1].alignment=CEN
wsk.column_dimensions["A"].width=36; wsk.column_dimensions["B"].width=28

# portada/orden de hojas: Indicadores primero
wb.move_sheet("Indicadores", -(len(wb.sheetnames)-1))

OUT="/root/alejo-psico/articulo/CliniA_export_sesiones.xlsx"
wb.save(OUT)
print("Guardado:", OUT)
print("Hojas:", wb.sheetnames)
print("Sesiones:", len(R))
